import pandas as pd
import openpyxl
import xlrd
import configparser


# Read the data from execel sheet
# dr= pd.read_excel("C:\\Users\\ADMIN\\Documents\\Login.xlsx")
# print(dr)

# read the specfic cell value from excel sheet:-->

# def read_specific_cell(file_path, sheet_name, target_value):
#     workbook =xopen_workbook_xls(ile_path)
#     # sheet = workbook.get(sheet_name)
#     # for sheets in sheet:
#     sh=workbook.get(sheet_name)
#     row=sh.nrows
#     for rownum in range(1,row):
#         print(sh.cell(rownum,2))
    # cell = sheet[target_value]
    # text=pd.read_excel(cell)
    # workbook.close
    # return text


file_path = "C:\\Users\\ADMIN\\Documents\\Login.xlsx"
sheet_name = "Sheet1"
target_value = "A1"

read_specific_cell("C:\\Users\\ADMIN\\Documents\\Login.xlsx",sheet_name,target_value)

# cell_coordinates=get_cell_coordinates(openpyxl.load_workbook(file_path))
# data=read_specific_cell(openpyxl.load_workbook(file_path),sheet_name,target_value)
# print(data)


# Write the data in the execl sheet
# data={"username":["HARI","RAMAN"],
#       "password":["RAT@1","HAT@2"]}
# obj=pd.DataFrame(data)
# obj.to_excel("C:\\Users\\ADMIN\\Documents\\Credentials.xlsx",index=False)


#Write the data for multiple sheets:-->

from xlsxwriter import Workbook


# data={ "Username" :['SATYAM','ABHI','Aditya','Rashmi'],
#       "Password":['SATYAM@1244','ABHI@2344','Aditiya@321','RASHMI@23445']}
#
#
# obj=pd.DataFrame(data)
# data1={ "Username" :['SATYAM','ABHI','Aditya','Rashmi'],
#       "Password":['SATYAM@1244','ABHI@2344','Aditiya@321','RASHMI@23445']}
#
#
# obj1=pd.DataFrame(data1)
# with pd.ExcelWriter('C:\\Users\\ADMIN\\Documents\\Weekend.xlsx',engine='xlsxwriter') as writer:
#    obj.to_excel(writer,sheet_name='Sheet1',index=False)
#    obj1.to_excel(writer, sheet_name='Sheet2', index=False)
# Read the data from properties file:-->

# def readData(file_path):
#     con=configparser.ConfigParser()
#     con.read(file_path)
#     return  con
# file_path="C:\\Users\\ADMIN\\PycharmProjects\\PTes\\Config.ini"
# properties=readData(file_path)
# username=properties.get("Login","username")
# print(username)
# password=properties.get("Login","password")
# print(password)
# url=properties.get("Login","url")
# print(url)


# Write data in properties file
# def writeData(file_path, properties):
#     con = configparser.ConfigParser()
#     con.read_dict(properties)
#     with open(file_path, "w") as configfile:
#         con.write(configfile)
#
#
# pro = {"Cre":{"username": "SATYAM",
#        "password": "TRSET@1",
#        "url": "www.blazedemo.com"}}
#
# writeData("TEST.ini", pro)
