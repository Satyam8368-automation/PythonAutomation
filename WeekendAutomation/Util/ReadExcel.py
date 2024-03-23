import pandas as pd
import openpyxl
from selenium  import webdriver

# d=pd.read_excel("C:\\Users\\ADMIN\\Documents\\Demo.xlsx")
#
# print(d)

# Write the data in the excel sheet:-->
#
# data={"username" :['TEST','REST','KICK'],
#       "password":['RTET','TSTSGSU','shgysg']}
#
# obj=pd.DataFrame(data)
# obj.to_excel("C:\\Users\\ADMIN\\Documents\\TESTRDAG.xlsx",index=False)

# Read the porperties file

# import configparser
#
# from selenium.webdriver.common.by import By
#
#
# def readdata(file_path):
#      d=configparser.ConfigParser()
#      d.read(file_path)
#      return d
#
# d_file_path="DATA.ini"
# properties=readdata(d_file_path)
# options=webdriver.ChromeOptions()
# options.add_experimental_option("detach",True)
# driver=webdriver.Chrome(options=options)
# driver.get(properties.get("Login","url"))
# driver.find_element(By.ID,properties.get("locators","id")).send_keys(properties.get("Login","username"))
# driver.find_element(By.ID,properties.get("locators","password")).send_keys(properties.get("Login","password"))
# driver.find_element(By.XPATH,properties.get("locators","button")).click()
# print(properties.get("Login","username"))
# print(properties.get("Login","password"))

import openpyxl
import pyautogui


#create a funcation to get the cordinates of a cell based on its values:-->


# def get_cell_coordinates(sheet,target_value):
#    for row in sheet.iter_rows():
#        for cell in row:
#            if cell.value== target_value:
#                return cell.coordinates


#to read the data


# def read_cell_value(file_path,sheet_name,cell_coordinates):
#    workbook=openpyxl.load_workbook(file_path)
#    sheet=workbook[sheet_name]
#    cell_value=sheet[cell_coordinates].value
#    # workbook.close()
#    return cell_value
#
#
# file_path="C:\\Users\\ADMIN\\Documents\\Demo.xlsx"
# sheet_name="Sheet1"
# target_value="A1"
#
#
# # cell_coordinates=get_cell_coordinates(sheet,target_value)
# # print(cell_coordinates)
# data=read_cell_value(openpyxl.load_workbook("C:\\Users\\ADMIN\\Documents\\Demo.xlsx"),sheet_name,target_value)
# print(data)


# import pandas as pd
# from xlsxwriter import Workbook
#
#
# data={ "Username" :['SATYAM','ABHI','Aditya','Rashmi'],
#       "Password":['SATYAM@1244','ABHI@2344','Aditiya@321','RASHMI@23445']}
#
#
# obj=pd.DataFrame(data)
# data1={ "Username" :['SATYAM','ABHI','Aditya','Rashmi'],
#       "Password":['SATYAM@1244','ABHI@2344','Aditiya@321','RASHMI@23445']}
#
#
# obj1=pd.DataFrame(data1)
# with pd.ExcelWriter('C:\\Users\\ADMIN\\Documents\\EW.xlsx',engine='xlsxwriter') as writer:
#    obj.to_excel(writer,sheet_name='Sheet1',index=False)
#    obj1.to_excel(writer, sheet_name='Sheet2', index=False)

import configparser




def write_properties(file_path,properties):
   config=configparser.ConfigParser()
   config.read_dict(properties)
   with open(file_path,'w') as configfile:
       config.write(configfile)


pro={
   'Settings':{
       'username':'Satyam@1',
       'password':'SATYA@12',
       'url':'https://blazedemo.com/login'
   }
}
write_properties('SAT.ini',pro)
